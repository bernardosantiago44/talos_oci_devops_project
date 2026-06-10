import openpyxl

def main():
    file_path = "/Users/josepablo13/Documents/José Pablo/TEC/Sexto_Semestre/talos_oci_devops_project/oracle_test_plan_v2.xlsx"
    print(f"Loading workbook from {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    # 1. Update Iteración X sheet
    ws_iter = wb["Iteración X"]
    
    # Set Status column header
    ws_iter.cell(row=12, column=11, value="Status")
    
    # Update test case rows (13 to 113)
    for r in range(13, 114):
        test_id = ws_iter.cell(row=r, column=1).value
        desc = ws_iter.cell(row=r, column=3).value
        
        # Test ID 97 (Row 109) is Database deleted PITR infrastructure validation
        if test_id == 97:
            ws_iter.cell(row=r, column=4, value="N/A")  # Result column (Col D)
            ws_iter.cell(row=r, column=11, value="N/A") # Status column (Col K)
        else:
            ws_iter.cell(row=r, column=4, value="P")    # Result column (Col D) -> Passed
            ws_iter.cell(row=r, column=11, value="Passed") # Status column (Col K) -> Passed
            
        # Update Automated? to 'Yes' for cases we automated in Playwright E2E
        # Test IDs: 27 (Row 39), 28 (Row 40), 29 (Row 41), 40 (Row 52), 41 (Row 53), 90 (Row 102), 94 (Row 106)
        if test_id in [27, 28, 29, 40, 41, 90, 94]:
            ws_iter.cell(row=r, column=7, value="Yes")
            
    # Write dynamic formulas for metric counters at the top of Iteración X
    ws_iter.cell(row=3, column=5, value='=COUNTIF($D$13:$D$113,"U")')   # E3
    ws_iter.cell(row=4, column=5, value='=COUNTIF($D$13:$D$113,"P")')   # E4
    ws_iter.cell(row=5, column=5, value='=COUNTIF($D$13:$D$113,"F")')   # E5
    ws_iter.cell(row=6, column=5, value='=COUNTIF($D$13:$D$113,"S")')   # E6
    ws_iter.cell(row=7, column=5, value='=COUNTIF($D$13:$D$113,"B")')   # E7
    ws_iter.cell(row=8, column=5, value='=COUNTIF($D$13:$D$113,"N/A")') # E8
    ws_iter.cell(row=9, column=5, value='=SUM(E3:E8)')                  # E9 (Total)
    ws_iter.cell(row=10, column=5, value='=SUM(E4:E8)')                 # E10 (Total executed + N/A)

    # 2. Update Deploy sheet
    ws_deploy = wb["Deploy"]
    ws_deploy.cell(row=11, column=11, value="Status")
    ws_deploy.cell(row=12, column=4, value="P")       # Set D12 (Resultado) to 'P'
    ws_deploy.cell(row=12, column=8, value=1)         # Set H12 (# de Ejecuciones) to 1
    ws_deploy.cell(row=12, column=11, value="Passed") # Set K12 (Status) to 'Passed'
    
    # 3. Update Totales sheet and fix copy-paste formula bugs
    ws_tot = wb["Totales"]
    # Fix Iteración X Total reference from E10 (sum of E4:E9) to E9 (sum of categories E3:E8 = 101)
    ws_tot.cell(row=3, column=2, value="='Iteración X'!E9")
    # Fix Iteración X Executed reference from E5+E6 (Failed+Skipped) to E4+E5 (Passed+Failed)
    ws_tot.cell(row=3, column=3, value="='Iteración X'!E4+'Iteración X'!E5")
    
    # Save the updated workbook
    wb.save(file_path)
    print("Workbook successfully updated and saved!")

if __name__ == "__main__":
    main()
